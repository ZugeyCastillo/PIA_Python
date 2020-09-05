import pandas as pd
import xlsxwriter
from clases2 import Cliente
from clases2 import Prestamo
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta

class PrestamosExcel:
    # este metodo nos permite leer los prestamos de la hoja Prestamos del excel
    def LeerPrestamos(self):
        # se definen las columnas del excel que se quieren ver, si solo se ponen dos pandas nos muestra dos
        # para este caso definimos todas las columnas que se encuentran en la hoja clientes del excel banco
        # tambien debe definirse el tipo de dato para cada columna
        columnas = {"IdCliente":int,"Monto":float,"FechaCreacion":str,"FechaExpiracion":str,"TipoPrestamo":int,"AbonoTotal":float,"Pagado":int,"NombreCliente":str}

        # hoja que pandas va a ir a consultar
        hoja = "Prestamos"

        # se utiliza la funcion read_excel de pandas la cual recibe como parametros el nombre del excel,
        # la hoja y las columnas que definimos arriba. Como resultado nos regresa un data frame
        df = pd.read_excel("Banco2.xlsx",sheet_name=hoja,dtype=columnas)
        return df

    # este metodo nos permite leer la lista de intereses de la hoja Intereses del excel
    def LeerIntereses(self):
        columnas = {"Tipo":int,"Duracion":str,"Intereses":str}

        # hoja que pandas va a ir a consultar
        hoja = "Intereses"

        # se utiliza la funcion read_excel de pandas la cual recibe como parametros el nombre del excel,
        # la hoja y las columnas que definimos arriba. Como resultado nos regresa un data frame
        df = pd.read_excel("Banco2.xlsx",sheet_name=hoja,dtype=columnas)
        return df

    # este metodo nos ayuda a leer los prestamos activos por cliente, si el data frame que regresa tiene valor
    # quiere decir que ese cliente tiene prestamos activos, si el data frame esta vacio entonces el cliente no tiene prestamos activos
    def LeerPrestamosActivosCliente(self, clienteId):
        df = self.LeerPrestamos()

        # IdCliente == ' + str(clienteId) filtra todos los registros del excel en cuya columna IdCliente tengan
        # valor igual al clienteId que recibimos como parametro
        # & Pagado == 0 filtra los registros del excel en cuya columna Pagado tengan un valor igual a 0, es decir
        # prestamos que aun no han sido pagados
        newDF = df.query('IdCliente == ' + str(clienteId) + ' & Pagado == 0')

        # si el data frame tiene valores significa que ese cliente tiene un prestamo activo
        # si el data frame esta vacio (empty) significa que el cliente no tiene prestamos activos
        return newDF

    def LeerAdeudoCliente(self, clienteId):

        if self.RevisarExpirado(clienteId):
            print("Su prestamo expiró y se realizó una reestructuración\n")

        df = self.LeerPrestamosActivosCliente(clienteId)
        filt = (df['IdCliente'] == clienteId)
        adeudo = round(float(df.loc[filt, 'Monto']), 2) - round(float(df.loc[filt, 'AbonoTotal']), 2)

        return adeudo

    # regresa todos los prestamos de un cliente pagados y no pagados
    def LeerPrestamosCliente(self, clienteId):
        # nos regresa tooooodos los prestamos
        df = self.LeerPrestamos()

        # filtramos los registros que tengan en la columna IdCliente un valor igual a clienteId
        newDF = df.query('IdCliente == ' + str(clienteId))
        return newDF

    def LeerPrestamosActivos(self):
        # nos regresa tooooodos los prestamos
        df = self.LeerPrestamos()

        # filtramos los registros que tengan en la columna Pagado un valor igual a 0, es decir
        # prestamos que aun no han sido pagados
        newDF = df.query('Pagado == 0')
        return newDF

    def LeerPrestamosPagados(self):
        # nos regresa tooooodos los prestamos
        df = self.LeerPrestamos()

        # filtramos los registros que tengan en la columna Pagado un valor igual a 1, es decir
        # prestamos que ya han sido pagados
        newDF = df.query('Pagado == 1')
        return newDF

    def LeerPrestamosFecha(self, fechaInicio, fechaFin):
        # se reciben las fechas como string en formato 2020-10-01
        # se convierten las fechas string a fecha objeto
        fechaI = datetime.strptime(fechaInicio, '%Y-%m-%d')
        fechaF = datetime.strptime(fechaFin, '%Y-%m-%d')

        # leemos todos los prestamos
        df = self.LeerPrestamos()

        # convertimos los valores de la columna FechaCreacion a objeto para poder comparar
        # con las fechas que recibimos como parametros
        df['FechaCreacion'] = pd.to_datetime(df['FechaCreacion'], format='%Y-%m-%d')

        # creamos un nuevo data frame que va a contener solo los registros que entran en el rango de fechas
        filt = df[(df['FechaCreacion'] >= fechaI) & (df['FechaCreacion'] <= fechaF)] 
        
        return filt

    def AgregarPrestamo(self, prestamo):
        df = self.LeerPrestamos()

        workbook = load_workbook('Banco2.xlsx')
        writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        newDF = df.append(prestamo.__dict__, ignore_index=True)
        newDF.to_excel(writer, sheet_name="Prestamos", index=False)
        writer.save()
        writer.close()
    
    def AbonarPrestamo(self, clienteId, abono):

        df = self.LeerPrestamos()
        filt = (df['IdCliente'] == clienteId) & (df['Pagado'] == 0)
        #filt = df[(df['FechaCreacion'] >= fechaI) & (df['FechaCreacion'] <= fechaF)] 

        abonoTotal = round(float(df.loc[filt, 'AbonoTotal']),2)
        monto = round(float(df.loc[filt, 'Monto']), 2)
        cambio = 0

        if (abonoTotal + abono) >= monto:
            cambio = (abonoTotal + abono) - monto
            df.loc[filt, 'AbonoTotal'] = (abonoTotal + abono) - cambio
            df.loc[filt, 'Pagado'] = 1
        else:
            df.loc[filt, 'AbonoTotal'] = abonoTotal + abono

        workbook = load_workbook('Banco2.xlsx')
        writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name="Prestamos", index=False)
        writer.save()
        writer.close()

        return cambio
    
    def RevisarExpirado(self, clienteId):
        # leemos todos los prestamos
        df = self.LeerPrestamos()
        dfOld = self.LeerPrestamos()

        today = date.today()
        
        # convertimos los valores de la columna FechaCreacion a objeto para poder comparar
        # con las fechas que recibimos como parametros
        dfOld['FechaExpiracion'] = pd.to_datetime(dfOld['FechaExpiracion'], format='%Y-%m-%d')

        resultados = dfOld[(dfOld['IdCliente'] == clienteId) & (dfOld['Pagado'] == 0) & (dfOld['FechaExpiracion'] < pd.to_datetime(today))] 

        filt = (df['IdCliente'] == clienteId) & (df['Pagado'] == 0)

        if not resultados.empty:
            nuevoMonto = round(float(df.loc[filt, 'Monto']), 2) - round(float(df.loc[filt, 'AbonoTotal']),2)
            nuevoMontoConImpuestos = nuevoMonto * 1.3
            nuevoTipo = 3
            df.loc[filt, 'Monto'] = nuevoMontoConImpuestos
            df.loc[filt, 'AbonoTotal'] = 0
            df.loc[filt, 'TipoPrestamo'] = nuevoTipo
            df.loc[filt, 'FechaExpiracion'] = str(date.today() + relativedelta(months=+12))

            workbook = load_workbook('Banco2.xlsx')
            writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            df.to_excel(writer, sheet_name="Prestamos", index=False)
            writer.save()
            writer.close()

            return True
        else:
            return False