import pandas as pd
import xlsxwriter
from clases2 import Cliente
from openpyxl import load_workbook
import numpy

class ClientesExcel:
    # regresa toda la lista de clientes
    def LeerClientes(self):
        # se definen las columnas del excel que se quieren ver, si solo se ponen dos pandas nos muestra dos
        # para este caso definimos todas las columnas que se encuentran en la hoja clientes del excel banco
        # tambien debe definirse el tipo de dato para cada columna
        columnas = {"Id":int,"\nNombre":str,"\nCorreo":str,"\nTelefono":str,"\nDireccion":str,"\nMonto":float,"\nEstatus":int}

        # hoja que pandas va a ir a consultar
        hoja = "Clientes"

        # se utiliza la funcion read_excel de pandas la cual recibe como parametros el nombre del excel,
        # la hoja y las columnas que definimos arriba. Como resultado nos regresa un data frame
        df = pd.read_excel("Banco2.xlsx",sheet_name=hoja,dtype=columnas)
        return df

    def AgregarCliente(self, cliente):
        # consultamos la funcion LeerClientes para obtener el data frame
        # posteriormente le agregaremos el nuevo cliente y guardaremos nuevamente el data frame en el excel
        df = self.LeerClientes()

        # vemos cual es el ultimo index que hay en data frame utilizando la funcion last_valid_index
        # se valida si ultimoId tiene valor, porque si no entonces se inicializa como -1 ya que quiere decir
        # que no hay registros en el excel
        ultimoId = df.last_valid_index()
        if ultimoId is None:
            ultimoId = -1

        # la funcion nos retorna el index del ultimo registro y le sumamos 1 para este asignarlo 
        # como id del objeto cliente que recibimos como parametro
        cliente.Id = ultimoId + 1

        # definimos el estatus del cliente como 1 (activo)
        # cuando eliminamos un cliente, este nunca se borra del excel, solo se cambia su estatus a 0 (inactivo)
        cliente.Estatus = 1

        # el data frame obtenido (df) es inmutable, es decir que no puede ser editado
        # es por eso que para agregar un cliente creamos un data frame nuevo llamado newDF el cual es igual
        # al data frame consultado agregandole con la funcion append el nuevo cliente
        # la funcion append del data frame recibe como parametro un diccionario, es por eso que convertimos el
        # objeto cliente a diccionario utilizando cliente.__dict__ mientras que ignore_index=True nos ayuda a
        # especificar que queremos que el index que asigne pandas al registro sea el consecutivo al ultimo registro
        newDF = df.append(cliente.__dict__, ignore_index=True)

        # utilizando la libreria openpyxl creamos un objeto workbook para especificar el excel que estaremos utilizando
        workbook = load_workbook('Banco2.xlsx')
        # creamos una variable a partir de la funcion de pandas ExcelWriter que sirve para definir que excel
        # vamos a manipular asi como que motor para editar el excel vamos a utilizar, utilizaremos openpyxl para este caso
        writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
        # las siguientes dos lineas sirven para que al guardar, tambien se guarden las otras hojas del archivo que no tuvieron cambios
        # lo hicimos asi porque de otra forma cuando guardabamos en una hoja, las demas hojas como Prestamos o Intereses
        # se borraban del excel
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)

        # utilizamos la funcion to_excel del data frame que ya contiene el cliente que agregamos y que nos permitira
        # unir la variable writer al data frame que estamos utilizando. Le especificamos que writer vamos a utilizar 
        # asi como que hoja del excel queremos modificar y que no queremos especificar un index
        newDF.to_excel(writer, sheet_name="Clientes", index=False)
        # finalmente utilizamos la funcion save del writer para guardar los cambios realizados al excel
        writer.save()
        writer.close()

    # se leen los clientes por nombre
    def LeerClientesNombre(self, nombre):
        # se utiliza la funcion que regresa el data frame con todos los clientes del excel para luego aplicarle
        # un filtro de pandas para obtener los que coinciden con el nombre
        df = self.LeerClientes()

        # se crea la variable filt la cual es un filtro. Se especifica utilizando corchetes sobre que columna del
        # se hara el filtro, luego se convierte a string utilizando .str y finalmente se utiliza .contains para
        # buscar los registros que contienen el nombre que buscamos en la columna nombre.
        filt = (df['Nombre'].str.contains(nombre))

        # se retorna el data frame, aplicandole el filtro que se creo anteriormente
        return df[filt]

    def LeerClientesId(self, id):
        # se utiliza la funcion que regresa el data frame con todos los clientes del excel para luego aplicarle
        # un filtro de pandas para obtener los que coinciden con el nombre
        df = self.LeerClientes()

        # se crea un filtro el cual busca los registros cuya columna Id sea igual al id que recibimos como parametro
        filt = (df['Id'] == int(id))
        # se retorna el data frame, aplicandole el filtro que se creo anteriormente
        return df[filt]

    def LeerClientesActivos(self):
        df = self.LeerClientes()

        filt = (df['Estatus'] == 1)
        return df[filt]
    
    def LeerClientesInactivos(self):
        df = self.LeerClientes()

        filt = (df['Estatus'] == 0)
        return df[filt]
 
    def LeerClienteObjeto(self, id):
        # se utiliza la funcion que regresa el data frame con todos los clientes del excel para luego aplicarle
        # un filtro de pandas para obtener los que coinciden con el nombre
        df = self.LeerClientes()

        # se crea un objeto cliente
        cliente = Cliente()

        # el data frame de pandas en un arreglo el cual tiene un index que es igual al id del cliente.
        # cuando creamos un nuevo cliente, obtenemos el ultimo index del data frame y se le suma uno para
        # darle este nuevo valor al id del cliente.
        # con la funcion del data frame .loc podemos obtener un registro a partir de su index.
        # aqui estamos buscando un cliente cuyo index sea igual al id que estamos pasandole y este lo asignamos
        # al objeto cliente que tenemos
        cliente = df.loc[int(id)]

        # finalmente retornamos el objeto cliente
        return cliente

    def AcualizarCliente(self, id, propiedad, nuevoValor):
        # se utiliza la funcion que regresa el data frame con todos los clientes del excel para luego aplicarle
        # un filtro de pandas para obtener los que coinciden con el nombre
        df = self.LeerClientes()

        # con la funcion .loc del data frame tambien podemos obtener un valor especifico de un registro y/o asignarle un nuevo valor
        # en este caso especificamos dos cosas, el index del registro (que es igual al id que tiene el cliente) y
        # la columna que queremos obtener, aqui un ejemplo df.loc[0, 'Nombre'] = 'nuevoNombre'
        # en el ejemplo queremos obtener la columna "Nombre" del registro con index = 0 (ese 0 es el id del cliente)
        # y le asignamos un nuevo valor con = 'nuevoNombre'
        df.loc[int(id), propiedad] = nuevoValor


        workbook = load_workbook('Banco2.xlsx')
        writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name="Clientes", index=False)
        writer.save()
        writer.close()

    def EliminarCliente(self, id):
        df = self.LeerClientes()

        df.loc[int(id), "Estatus"] = 0

        workbook = load_workbook('Banco2.xlsx')
        writer = pd.ExcelWriter('Banco2.xlsx', engine="openpyxl")
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet_name="Clientes", index=False)
        writer.save()
        writer.close()
